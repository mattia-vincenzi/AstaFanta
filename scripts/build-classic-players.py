import json
import sys
from openpyxl import load_workbook

source, target = sys.argv[1:3]
workbook = load_workbook(source, read_only=True, data_only=True)
sheet = workbook['Tutti']
players = []
for row in sheet.iter_rows(min_row=3, values_only=True):
    if not row[0] or not row[1] or not row[3]:
        continue
    players.append({'id': str(row[0]), 'roles': [str(row[1]).strip()], 'name': str(row[3]).strip(), 'team': str(row[4]).strip(), 'qt': float(row[5] or 0), 'fvm': float(row[11] or 0)})
workbook.close()
with open(target, 'w', encoding='utf-8') as handle:
    json.dump(players, handle, ensure_ascii=False, separators=(',', ':'))
